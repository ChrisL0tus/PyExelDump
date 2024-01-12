from openpyxl import load_workbook

workbook = load_workbook(filename='./refer/past/14.xlsx')
print(workbook.sheetnames)
sheet = workbook['Sheet1']
print(sheet)
print(sheet.dimensions)

cells = sheet[sheet.dimensions]

with open('output14.txt', 'w', encoding='utf-8') as file:
    for cell_rows in cells:
        for cell_columns in cell_rows:
            value = cell_columns.value
            if value is not None:
                file.write(str(value) + '\n')
